import streamlit as st
import openpyxl
from openpyxl.styles import Protection
import tempfile

def lock_and_hide_formulas(input_file_path, output_file_path):
    # Baca file Excel yang sudah ada
    workbook = openpyxl.load_workbook(input_file_path)

    # Loop melalui semua lembar kerja dalam file
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Kunci seluruh sel di sheet dan sembunyikan rumus
        for row in sheet.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True, hidden=True)

        # Buka kunci sel dari kolom E hingga P dan baris 1 hingga 1000, dan tampilkan rumus
        for row in sheet.iter_rows(min_row=1, max_row=1000, min_col=5, max_col=16):
            for cell in row:
                cell.protection = Protection(locked=False, hidden=False)

        # Aktifkan proteksi lembar kerja dengan password
        sheet.protection.sheet = True
        sheet.protection.set_password('dsinternational')  # Password untuk proteksi

    # Simpan perubahan
    workbook.save(output_file_path)

st.title('Excel Lock & Hide Formulas')

uploaded_file = st.file_uploader('Upload your Excel file', type='xlsx')

if uploaded_file is not None:
    # Buat file sementara untuk menyimpan file hasil
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        input_file_path = tmp.name
        tmp.write(uploaded_file.read())
        tmp.seek(0)
        output_file_path = input_file_path.replace('.xlsx', '_locked_hidden.xlsx')
        
        # Proses file
        lock_and_hide_formulas(input_file_path, output_file_path)
        
        # Baca kembali file hasil untuk di-download oleh pengguna
        with open(output_file_path, 'rb') as f:
            st.download_button(
                label="Download processed file",
                data=f,
                file_name="locked_hidden_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
